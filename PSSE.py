import os,sys
import psse35
import psspy
import math
from library import openpyxl
from psspy import _i,_f,_s
from openpyxl import load_workbook
file_path=r"C:\Users\Thien Dang\Desktop\ins\training\training\13_Water_Spring_Solar\Water_Spring_Solar.xlsm"
case_path=r"C:\Users\Thien Dang\Desktop\ins\code\test.sav"
# sys_path_PSSE=r'C:\Program Files\PTI\PSSE35\35.3\PSSPY27'
# sys.path.append(sys_path_PSSE)
# os_path_PSSE=r'C:\Program Files\PTI\PSSE35\35.3\PSSBIN'
# os.environ['PATH']+=";"+os_path_PSSE
# os.environ['PATH']+=";"+sys_path_PSSE

ierr=psspy.psseinit(50)
ierr=psspy.newcase_2(basemva=100, basefreq=50)
psspy.newdiagfile()

data_excel=openpyxl.load_workbook(file_path,data_only=True)

def sld(res):
    mpt = len(res)
    len1=[]
    len2 = [0] * (max(res) + 1)
    dem = [0] * (max(res) + 1)
    psspy.bus_data_4(999997,0,[1,1,1,1],[0.0,1.0,0.0,1.1,0.9,1.1,0.9],r"""DUMMY_POI""")
    for i in range(1,mpt+1):
        len1.append(-i*2)
        name = f"GSU{i}_SEC"
        sld =f"BU {10000*i + 1}"
        frombus =10000*i+1
        psspy.bus_data_4(frombus,0,[2,1,1,1],[0.0,1.0,0.0,1.1,0.9,1.1,0.9],name)
        ## Gen
        psspy.plant_data_4(frombus,0,[999997,0],[1,100.0])
        psspy.machine_data_4(frombus,r"""1""",[1,1,0,0,0,1,0],[0.0,0.0,9999.0,-9999.0,9999.0,-9999.0,100.0,0.0,1.0,0.0,0.0,1.0,1.0,1.0,1.0,1.0,1.0],"")

        name1 = f"GSU{i}_PRI"
        sld1 =f"BU {10000+1000*i}"
        tobus =10000+1000*i
        psspy.bus_data_4(tobus,0,[1,1,1,1],[0.0,1.0,0.0,1.1,0.9,1.1,0.9],name1)
        psspy.growdiagram_2(1,1,[sld1],3,-i*2,[0,0,0,0,0,0,0,0,0,0,0,0,0])
        psspy.two_winding_data_6(tobus,frombus,r"""1""",[1,tobus,1,0,0,0,5,0,tobus,0,0,1,0,1,1,1],[0.0,0.0001,100.0,1.0,0.0,0.0,1.0,0.0,1.0,1.0,1.0,1.0,0.0,0.0,1.05,0.95,1.1,0.9,0.0,0.0,0.0],[0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,
        0.0,0.0,0.0],"","")
        psspy.growdiagram_2(1,1,[sld],1,-i*2,[0,0,0,0,0,0,0,0,0,0,0,0,0])
    for i in range(len(res)):
        len2[res[i]]=len2[res[i]]+len1[i]
        dem[res[i]]=dem[res[i]]+1
    for i in range(1,len(len2)):
        len2[i]=len2[i]/dem[i]
    position_110002=-(len(res)+1)
    psspy.growdiagram_2(1,1,[r"""BU 999997"""],13,position_110002,[0,0,0,0,0,0,0,0,0,0,0,0,0]) # vi position_110002 
    #creat MPT bus
    mpt_number=[]
    for i in range(1,max(res)+1):
        name = f"MPT{i}_SEC"
        sld =f"BU {100000 + 10000*i}"
        sec =100000 + 10000*i
        psspy.bus_data_4(sec,0,[1,1,1,1],[0.0,1.0,0.0,1.1,0.9,1.1,0.9],name)
        psspy.load_data_6(sec,r"""1""",[1,1,1,1,1,0,0],[0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0],"")
        psspy.switched_shunt_data_5(sec,r"""1""",[_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s)
        #thay doi vi tri bus 110000
        psspy.growdiagram_2(1,1,[sld],5,len2[i],[0,0,0,0,0,0,0,0,0,0,0,0,0])
        mpt_number.append(sec)

        name2 = f"MPT{i}_TER"
        sld2 =f"BU {100000 + 10000*i+3}"
        ter =100000 + 10000*i+3
        psspy.bus_data_4(ter,0,[1,1,1,1],[0.0,1.0,0.0,1.1,0.9,1.1,0.9],name2)
        psspy.growdiagram_2(1,1,[sld2],6,len2[i]+1,[0,0,0,0,0,0,0,0,0,0,0,0,0])

        if i==1:
            psspy.bus_data_4(110002,0,[1,1,1,1],[0.0,1.0,0.0,1.1,0.9,1.1,0.9],r"""MPT1_PRI""")
            psspy.growdiagram_2(1,1,[r"""BU 110002"""],9,position_110002,[0,0,0,0,0,0,0,0,0,0,0,0,0])

        ## three wnd
        psspy.three_wnd_imped_data_4(110002,sec,ter,r"""1""",[_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,1],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],"","")
        psspy.three_wnd_winding_data_5(110002,sec,ter,r"""1""",1,[_i,_i,_i,_i,_i,1],[_f,_f,_f,_f,_f,1.02,0.98,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
        psspy.three_wnd_winding_data_5(110002,sec,ter,r"""1""",1,[_i,_i,_i,_i,_i,-1],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
        psspy.three_wnd_winding_data_5(110002,sec,ter,r"""1""",2,[_i,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,1.02,0.98,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
        psspy.three_wnd_winding_data_5(110002,sec,ter,r"""1""",3,[_i,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,1.02,0.98,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
        psspy.growdiagram_2(1,1,[sld],5,-i*2,[0,0,0,0,0,0,0,0,0,0,0,0,0])

    ##creat branch
    for i in range(mpt):

        frombus = 10000+1000*(i+1)
        sld =f"BU {frombus}"
        tobus = mpt_number[res[i]-1]
        psspy.branch_data_3(frombus,tobus,r"""1""",[1,frombus,1,0,0,0],[0.0,0.0001,0.0,0.0,0.0,0.0,0.0,0.0,1.0,1.0,1.0,1.0],[0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0],"")
        psspy.growdiagram_2(1,1,[sld],1,-i*2,[0,0,0,0,0,0,0,0,0,0,0,0,0])


    psspy.bus_data_4(960000,0,[1,1,1,1],[0.0,1.0,0.0,1.1,0.9,1.1,0.9],r"""DUMMY_SUB""")
    psspy.growdiagram_2(1,1,[r"""BU 960000"""],10,position_110002,[0,0,0,0,0,0,0,0,0,0,0,0,0])
    psspy.bus_data_4(970000,0,[3,1,1,1],[0.0,1.0,0.0,1.1,0.9,1.1,0.9],r"""POI""")
    psspy.growdiagram_2(1,1,[r"""BU 970000"""],15,position_110002,[0,0,0,0,0,0,0,0,0,0,0,0,0])
    psspy.branch_data_3(110002,960000,r"""1""",[1,960000,1,0,0,0],[0.0,0.0001,0.0,0.0,0.0,0.0,0.0,0.0,1.0,1.0,1.0,1.0],[0.0,0.0001,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0],"")
    psspy.growdiagram_2(1,1,[r"""BU 110002"""],1,position_110002,[0,0,0,0,0,0,0,0,0,0,0,0,0])

    psspy.branch_data_3(960000,999997,r"""1""",[1,999997,1,0,0,0],[0.0,0.0001,0.0,0.0,0.0,0.0,0.0,0.0,1.0,1.0,1.0,1.0],[0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0],"")
    psspy.growdiagram_2(1,1,[r"""BU 960000"""],1,position_110002,[0,0,0,0,0,0,0,0,0,0,0,0,0])

    psspy.plant_data_4(970000,0,[0,0],[1.0,100.0])
    psspy.machine_data_4(970000,r"""1""",[1,1,0,0,0,0,0],[0.0,0.0,9999.0,-9999.0,9999.0,-9999.0,100.0,0.0,1.0,0.0,0.0,1.0,1.0,1.0,1.0,1.0,1.0],"")

    psspy.branch_data_3(999997,970000,r"""1""",[1,970000,1,0,0,0],[0.0,0.0001,0.0,0.0,0.0,0.0,0.0,0.0,1.0,1.0,1.0,1.0],[0.0,0.0001,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0],"")
    psspy.growdiagram_2(1,1,[r"""BU 970000"""],1,position_110002,[0,0,0,0,0,0,0,0,0,0,0,0,0])

    psspy.newseq()
    psspy.seq_branch_data_3(110002,960000,r"""1""",_i,[_f,0.0001,_f,_f,_f,_f,_f,_f])
    psspy.seq_branch_data_3(970000,999997,r"""1""",_i,[_f,0.0001,_f,_f,_f,_f,_f,_f])

    psspy.machine_chng_4(970000,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,10000.0,_f,0.1,_f,_f,_f,_f,_f,_f,_f,_f],_s)
    psspy.seq_machine_data_4(970000,r"""1""",_i,[_f,0.01,_f,0.01,_f,0.01,0.01,0.01,_f,_f,_f])
    ### creat Transformer
    ierr=psspy.save(r'C:\Users\Thien Dang\Desktop\ins\code\test.sav')
    psspy.savediagfile(r'C:\Users\Thien Dang\Desktop\ins\code\test.sld')
def data_entry():
    #Nhập điện áp cho bus
    for i in range(1,dem[1]+1):
        sheet1=data_excel[f'1 General_{i}']
        v_gsu_pri = sheet1["B11"].value
        psspy.bus_chng_3(10000+i*1000,[_i,_i,_i,_i],[v_gsu_pri,_f,_f,_f,_f,_f,_f])
        v_gsu_sec = sheet1["B10"].value
        psspy.bus_chng_3(i*10000+1,[_i,_i,_i,_i],[v_gsu_sec,_f,_f,_f,_f,_f,_f])
        v_mpt_sec=sheet1['B11']
        psspy.bus_chng_3(100000+i*10000,[_i,_i,_i,_i],[v_gsu_pri,_f,_f,_f,_f,_f,_f])

    # Nhập gsu
    for i in range (1,dem[2]+1):
        sheet2=data_excel[f'2 XFMR Impedance_{i}']
        gsu_r=sheet2['B12'].value
        gsu_x=sheet2['B13'].value
        gsu_v1=sheet2['B5'].value
        gsu_v2=sheet2['C5'].value
        gsu_r0=sheet2['B34'].value
        gsu_x0=sheet2['B35'].value
        gsu_vectorgroup=sheet2['B3'].value
        psspy.two_winding_chng_6(10000*i+1,10000+1000*i,r"""1""",[_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,2,_i],[gsu_r,gsu_x,_f,_f,gsu_v1,_f,_f,gsu_v2,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s,r""f"{gsu_vectorgroup}""")
        psspy.seq_two_winding_data_3(10000*i+1,10000+1000*i,r"""1""",[14,2,2],[_f,_f,gsu_r0,gsu_x0,_f,_f,_f,_f,_f,_f])

    # Nhập branch 110002-970000
    v_110002=sheet1['B12'].value
    psspy.bus_chng_3(110002,[_i,_i,_i,_i],[v_110002,_f,_f,_f,_f,_f,_f])
    psspy.bus_chng_3(960000,[_i,_i,_i,_i],[v_110002,_f,_f,_f,_f,_f,_f])
    psspy.bus_chng_3(999997,[_i,_i,_i,_i],[v_110002,_f,_f,_f,_f,_f,_f])
    psspy.bus_chng_3(970000,[_i,_i,_i,_i],[v_110002,_f,_f,_f,_f,_f,_f])
    sheet3=data_excel['3 OH line impedance']
    ohline_r=sheet3['AD2'].value
    ohline_x=sheet3['AE2'].value
    ohline_b=sheet3['AJ2'].value
    ohline_r0=sheet3['AF2'].value
    ohline_x0=sheet3['AG2'].value
    ohline_b0=sheet3['AK2'].value
    ohline_mva1=sheet3['I2'].value
    ohline_mva2=ohline_mva1*1.15
    psspy.branch_chng_3(960000,999997,r"""1""",[_i,_i,_i,_i,_i,_i],[ohline_r,ohline_x,ohline_b,_f,_f,_f,_f,_f,_f,_f,_f,_f],[ohline_mva1,ohline_mva2,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s)
    psspy.seq_branch_data_3(960000,999997,r"""1""",_i,[ohline_r0,ohline_x0,ohline_b0,_f,_f,_f,_f,_f])

    # Nhập gen
    for i in range(1, dem[1]+1):
        sheet5=data_excel[f'5 Generator Impedance_{i}']
        r=sheet5['B13'].value
        x=sheet5['B14'].value
        x_subt=sheet5['B16'].value
        x_t=sheet5['B17'].value
        x_syn=sheet5['B18'].value
        r_neg=sheet5['B20'].value
        x_neg=sheet5['B21'].value
        r0=sheet5['B23'].value
        x0=sheet5['B24'].value
        rg=sheet5['B26'].value
        xg=sheet5['B27'].value
        sheet1=data_excel[f'1 General_{i}']
        mva=sheet1['B7'].value*sheet1['B9'].value
        gen_num=10000*i+1
        psspy.machine_chng_4(gen_num,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,mva,r,x,_f,_f,_f,_f,_f,_f,_f,_f],_s)
        psspy.seq_machine_data_4(gen_num,r"""1""",_i,[r,x_subt,r_neg,x_neg,r0,x0,x_t,x_syn,rg,xg,_f])
    # Nhập sheet 4 và mpt
    for l in range(1,dem[4]+1):
        sheet4=data_excel[f'4 UG collection sys impedance_{l}']
        for i in range(1,1000):
            mpt_number=0
            if sheet4[f'AH{i}'].value=='Equivalent R (pu)':
                for j in range (i+1,1000):
                    mpt_number = mpt_number+1
                    if isinstance(sheet4[f'AH{j}'].value, (int, float)):
                        r=sheet4[f'AH{(j)}'].value
                        x=sheet4[f'AI{(j)}'].value
                        b=sheet4[f'AJ{(j)}'].value
                        r0=sheet4[f'AM{(j)}'].value
                        x0=sheet4[f'AN{(j)}'].value
                        b0=sheet4[f'AO{(j)}'].value
                        rate1=sheet4[f'AT{(j)}'].value
                        rate2=sheet4[f'AU{(j)}'].value
                        psspy.branch_chng_3(10000+1000*l,100000+10000*mpt_number,r"""1""",[_i,_i,_i,_i,_i,_i],[r,x,b,_f,_f,_f,_f,_f,_f,_f,_f,_f],[rate1,rate2,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s)
                        psspy.seq_branch_data_3(10000+1000*l,100000+10000*mpt_number,r"""1""",_i,[r0,x0,b0,_f,_f,_f,_f,_f])
                        sheet2=data_excel[f'2 XFMR Impedance_{l}']
                        v_110003=sheet2['G5'].value
                        r12=sheet2['D12'].value
                        x12=sheet2['D13'].value
                        r23=sheet2['H12'].value
                        x23=sheet2['H13'].value
                        r13=sheet2['F12'].value
                        x13=sheet2['F13'].value
                        mva_base=sheet2['D14'].value
                        v_pri=sheet2['D5'].value
                        v_sec=sheet2['E5'].value
                        v_ter=sheet2['G5'].value
                        r120=sheet2['D34'].value
                        x120=sheet2['D35'].value
                        r230=sheet2['H34'].value
                        x230=sheet2['H35'].value
                        r130=sheet2['F34'].value
                        x130=sheet2['F35'].value
                        sec=100000+mpt_number*10000
                        pri=110002
                        ter=sec+3
                        psspy.bus_chng_3(ter,[_i,_i,_i,_i],[v_110003,_f,_f,_f,_f,_f,_f])
                        psspy.three_wnd_imped_chng_4(pri,sec,ter,r"""1""",[_i,_i,_i,_i,_i,2,2,_i,_i,_i,_i,_i,_i],[r12,x12,r23,x23,r13,x13,mva_base,mva_base,mva_base,_f,_f,_f,_f,_f,_f,_f,_f],_s,_s)
                        psspy.three_wnd_winding_data_5(pri,sec,ter,r"""1""",1,[_i,_i,sec,_i,_i,1],[_f,v_pri,_f,_f,_f,_f,_f,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
                        psspy.three_wnd_winding_data_5(pri,sec,ter,r"""1""",2,[_i,_i,_i,_i,_i,_i],[_f,v_sec,_f,_f,_f,_f,_f,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
                        psspy.three_wnd_winding_data_5(pri,sec,ter,r"""1""",3,[_i,_i,_i,_i,_i,_i],[_f,v_ter,_f,_f,_f,_f,_f,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
                        psspy.seq_three_winding_data_3(pri,sec,ter,r"""1""",[2,2,14],[_f,_f,r120,x120,_f,_f,r230,x230,_f,_f,r130,x130,_f,_f])
                        break 
    ierr=psspy.save(case_path)
dem = [0] * 7
for sheet_name in data_excel.sheetnames:
    for i in range(1,7):
        if sheet_name.startswith(f'{i}'):
            dem[i]+=1
a=[]
for l in range(1,dem[4]+1):
    sheet4=data_excel[f'4 UG collection sys impedance_{l}']
    for i in range(1,1000):
        mpt_number=0
        if sheet4[f'AH{i}'].value=='Equivalent R (pu)':
            for j in range (i+1,1000):
                mpt_number = mpt_number+1
                if isinstance(sheet4[f'AH{j}'].value, (int, float)):
                    a.append(mpt_number)
                    break
sld(a)
data_entry()