import os,sys
import psse35
import psspy
import numpy as np
import math
import openpyxl
from psspy import _i,_f,_s
from openpyxl import load_workbook
file_path=r"C:\Users\Thien Dang\Desktop\PVBESS1.xlsm"
case_path=r"C:\Users\Thien Dang\Desktop\ins\code\test.sav"
data={}
sys_path_PSSE=r'C:\Program Files\PTI\PSSE35\35.3\PSSPY27'
sys.path.append(sys_path_PSSE)
os_path_PSSE=r'C:\Program Files\PTI\PSSE35\35.3\PSSBIN'
os.environ['PATH']+=";"+os_path_PSSE
os.environ['PATH']+=";"+sys_path_PSSE
case=psspy.psseinit(50)
case=psspy.case(case_path)

data_excel=openpyxl.load_workbook(file_path,data_only=True)
dem = [0] * 7
for sheet_name in data_excel.sheetnames:
    for i in range(1,7):
        if sheet_name.startswith(f'{i}'):
            dem[i]+=1
# ierr, all_buses = psspy.abusint(string="NUMBER")
# ierr, bus_names = psspy.abuschar(string="NAME")
# for bus in bus_names:
#     print(bus)
#     if 'GSU' in bus and 'SEC' in bus:
#         dem2 +=1
# print(dem2)
# bus_dict = {bus  for bus in zip(all_buses[0])}

# Nhập điện áp cho bus
for i in range(1,dem[1]+1):
    sheet1=data_excel[f'1 General_{i}']
    v_gsu_pri = sheet1["B11"].value
    psspy.bus_chng_3(10000+i*1000,[_i,_i,_i,_i],[v_gsu_pri,_f,_f,_f,_f,_f,_f])
    v_gsu_sec = sheet1["B10"].value
    psspy.bus_chng_3(i*10000+1,[_i,_i,_i,_i],[v_gsu_sec,_f,_f,_f,_f,_f,_f])
    v_mpt_sec=sheet1['B11']
    psspy.bus_chng_3(100000+i*10000,[_i,_i,_i,_i],[v_gsu_pri,_f,_f,_f,_f,_f,_f])

# Nhập gsu
for i in range (1,dem[2]+1):
    sheet2=data_excel[f'2 XFMR Impedance_{i}']
    gsu_r=sheet2['B12'].value
    gsu_x=sheet2['B13'].value
    gsu_v1=sheet2['B5'].value
    gsu_v2=sheet2['C5'].value
    gsu_r0=sheet2['B34'].value
    gsu_x0=sheet2['B35'].value
    gsu_vectorgroup=sheet2['B3'].value
    psspy.two_winding_chng_6(10000*i+1,10000+1000*i,r"""1""",[_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,2,_i],[gsu_r,gsu_x,_f,_f,gsu_v1,_f,_f,gsu_v2,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s,r""f"{gsu_vectorgroup}""")
    psspy.seq_two_winding_data_3(10000*i+1,10000+1000*i,r"""1""",[14,2,2],[_f,_f,gsu_r0,gsu_x0,_f,_f,_f,_f,_f,_f])

# Nhập branch 110002-970000
v_110002=sheet1['B12'].value
psspy.bus_chng_3(110002,[_i,_i,_i,_i],[v_110002,_f,_f,_f,_f,_f,_f])
psspy.bus_chng_3(960000,[_i,_i,_i,_i],[v_110002,_f,_f,_f,_f,_f,_f])
psspy.bus_chng_3(999997,[_i,_i,_i,_i],[v_110002,_f,_f,_f,_f,_f,_f])
psspy.bus_chng_3(970000,[_i,_i,_i,_i],[v_110002,_f,_f,_f,_f,_f,_f])
sheet3=data_excel['3 OH line impedance']
ohline_r=sheet3['AD2'].value
ohline_x=sheet3['AE2'].value
ohline_b=sheet3['AJ2'].value
ohline_r0=sheet3['AF2'].value
ohline_x0=sheet3['AG2'].value
ohline_b0=sheet3['AK2'].value
ohline_mva1=sheet3['I2'].value
ohline_mva2=ohline_mva1*1.15
psspy.branch_chng_3(960000,999997,r"""1""",[_i,_i,_i,_i,_i,_i],[ohline_r,ohline_x,ohline_b,_f,_f,_f,_f,_f,_f,_f,_f,_f],[ohline_mva1,ohline_mva2,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s)
psspy.seq_branch_data_3(960000,999997,r"""1""",_i,[ohline_r0,ohline_x0,ohline_b0,_f,_f,_f,_f,_f])

# Đếm số mpt


# Nhập gen
for i in range(1, dem[1]+1):
    sheet5=data_excel[f'5 Generator Impedance_{i}']
    r=sheet5['B13'].value
    x=sheet5['B14'].value
    x_subt=sheet5['B16'].value
    x_t=sheet5['B17'].value
    x_syn=sheet5['B18'].value
    r_neg=sheet5['B20'].value
    x_neg=sheet5['B21'].value
    r0=sheet5['B23'].value
    x0=sheet5['B24'].value
    rg=sheet5['B26'].value
    xg=sheet5['B27'].value
    sheet1=data_excel[f'1 General_{i}']
    pgen=sheet1['B7'].value*sheet1['B9'].value
    gen_num=10000*i+1
    psspy.machine_chng_4(gen_num,r"""1""",[_i,_i,_i,_i,_i,_i,_i],[pgen,_f,_f,_f,_f,_f,_f,r,x,_f,_f,_f,_f,_f,_f,_f,_f],_s)
    psspy.seq_machine_data_4(gen_num,r"""1""",_i,[r,x_subt,r_neg,x_neg,r0,x0,x_t,x_syn,rg,xg,_f])
# Nhập sheet 4 và mpt
for l in range(1,dem[4]+1):
    sheet4=data_excel[f'4 UG collection sys impedance_{l}']
    for i in range(1,1000):
        mpt_number=0
        if sheet4[f'AH{i}'].value=='Equivalent R (pu)':
            for j in range (i+1,1000):
                mpt_number = mpt_number+1
                if isinstance(sheet4[f'AH{j}'].value, (int, float)):
                    r=sheet4[f'AH{(j)}'].value
                    x=sheet4[f'AI{(j)}'].value
                    b=sheet4[f'AJ{(j)}'].value
                    r0=sheet4[f'AM{(j)}'].value
                    x0=sheet4[f'AN{(j)}'].value
                    b0=sheet4[f'AO{(j)}'].value
                    rate1=sheet4[f'AT{(j)}'].value
                    rate2=sheet4[f'AU{(j)}'].value
                    psspy.branch_chng_3(10000+1000*l,100000+10000*mpt_number,r"""1""",[_i,_i,_i,_i,_i,_i],[r,x,b,_f,_f,_f,_f,_f,_f,_f,_f,_f],[rate1,rate2,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s)
                    psspy.seq_branch_data_3(10000+1000*l,100000+10000*mpt_number,r"""1""",_i,[r0,x0,b0,_f,_f,_f,_f,_f])
                    sheet2=data_excel[f'2 XFMR Impedance_{l}']
                    v_110003=sheet2['G5'].value
                    r12=sheet2['D12'].value
                    x12=sheet2['D13'].value
                    r23=sheet2['H12'].value
                    x23=sheet2['H13'].value
                    r13=sheet2['F12'].value
                    x13=sheet2['F13'].value
                    mva_base=sheet2['D14'].value
                    v_pri=sheet2['D5'].value
                    v_sec=sheet2['E5'].value
                    v_ter=sheet2['G5'].value
                    r120=sheet2['D34'].value
                    x120=sheet2['D35'].value
                    r230=sheet2['H34'].value
                    x230=sheet2['H35'].value
                    r130=sheet2['F34'].value
                    x130=sheet2['F35'].value
                    sec=100000+mpt_number*10000
                    pri=110002
                    ter=sec+3
                    psspy.bus_chng_3(ter,[_i,_i,_i,_i],[v_110003,_f,_f,_f,_f,_f,_f])
                    psspy.three_wnd_imped_chng_4(pri,sec,ter,r"""1""",[_i,_i,_i,_i,_i,2,2,_i,_i,_i,_i,_i,_i],[r12,x12,r23,x23,r13,x13,mva_base,mva_base,mva_base,_f,_f,_f,_f,_f,_f,_f,_f],_s,_s)
                    psspy.three_wnd_winding_data_5(pri,sec,ter,r"""1""",1,[_i,_i,sec,_i,_i,1],[_f,v_pri,_f,_f,_f,_f,_f,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
                    psspy.three_wnd_winding_data_5(pri,sec,ter,r"""1""",2,[_i,_i,_i,_i,_i,_i],[_f,v_sec,_f,_f,_f,_f,_f,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
                    psspy.three_wnd_winding_data_5(pri,sec,ter,r"""1""",3,[_i,_i,_i,_i,_i,_i],[_f,v_ter,_f,_f,_f,_f,_f,_f,_f,_f],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
                    psspy.seq_three_winding_data_3(pri,sec,ter,r"""1""",[2,2,14],[_f,_f,r120,x120,_f,_f,r230,x230,_f,_f,r130,x130,_f,_f])
                    break
    
ierr=psspy.save(r'C:\Users\Thien Dang\Desktop\ins\code\test.sav')